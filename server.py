from flask import Flask, request, render_template, send_file, session
from flask_session import Session
from src.predict import predict_sentiments
from src.youtube import get_video_comments
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from io import BytesIO
import datetime

app = Flask(__name__)
CORS(app)
app.config['SESSION_TYPE'] = 'filesystem'
Session(app)

# Global variable to store data
processed_data = {}

def get_video(video_id):
    if not video_id:
        return {"error": "video_id is required"}

    video_name, comments = get_video_comments(video_id)
    
    predictions = predict_sentiments(comments)

    positive = predictions.count("Positive")
    negative = predictions.count("Negative")

    summary = {
        "video_name": video_name,
        "positive": positive,
        "negative": negative,
        "num_comments": len(comments),
        "rating": (positive / len(comments)) * 100
    }

    return {"predictions": predictions, "comments": comments, "summary": summary}

@app.route('/', methods=['GET', 'POST'])
def index():
    global processed_data  # Use the global variable to store data

    summary = None      
    comments = []
    if request.method == 'POST':
        video_url = request.form.get('video_url')
        data = get_video(video_url)

        # Store the processed data in the global variable
        processed_data['summary'] = data['summary']
        processed_data['comments'] = list(zip(data['comments'], data['predictions']))

        summary = data['summary']
        comments = processed_data['comments']
        
        #session solotion
        session['summary'] = data['summary']
        session['comments'] = list(zip(data['comments'], data['predictions']))
        
    return render_template('index.html', summary=summary, comments=comments)

@app.route('/download_excel', methods=['POST'])
def download_excel():
    global processed_data  # Access the global variable to retrieve data
    
    summary = {}      
    comments = []

    # Retrieve the data from the global variable
    #summary = processed_data.get('summary')
    #comments = processed_data.get('comments')
    if "summary" in session:
     summary = session.get("summary", {})
     comments = session.get("comments", [])

    print (summary['video_name'])

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Comments Analysis"

    # Define a border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write the summary with borders
    summary_data = [
        ["video name", summary['video_name']],
        ["positive comments", summary['positive']],
        ["Negative Comments", summary['negative']],
        ["Total Comments", summary['num_comments']],
        ["Rating", f"{summary['rating']}%"]
    ]

    for row in summary_data:
        ws.append(row)
        for cell in ws[ws.max_row]:  # Apply border to the last appended row
            cell.border = thin_border

    ws.append([])  # Empty row for separation

    # Write the header for comments with borders
    header = ["Comment", "Sentiment"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.border = thin_border

    # Write the data with borders
    for comment, sentiment in comments:
        ws.append([comment, sentiment])
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Save the workbook to a BytesIO buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Get the current date for the filename
    today_date = datetime.datetime.today().strftime('%Y-%m-%d')
    base_filename = f"comments_analysis_{today_date}.xlsx"

    # Send the file to the client
    return send_file(output, download_name=base_filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)
